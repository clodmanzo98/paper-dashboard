import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NumberFormat
from openpyxl.utils import get_column_letter

# Carica il file
with open('paper_portfolio.json', 'r') as f:
    data = json.load(f)

# Prepariamo i DataFrame
df_history = pd.DataFrame(data['history'])
df_positions = pd.DataFrame(data['positions'])

# Pulizia e conversione date per lo storico
if not df_history.empty:
    df_history['data_apertura'] = pd.to_datetime(df_history['timestamp'], unit='s')
    df_history['data_chiusura'] = pd.to_datetime(df_history['exit_timestamp'], unit='s')
    # Riordino colonne
    cols = ['id', 'data_apertura', 'data_chiusura', 'market', 'side', 'entry_price', 'exit_price', 'cost', 'proceeds', 'pnl', 'pnl_pct', 'exit_reason', 'leader']
    df_history = df_history[cols]

# Pulizia date per posizioni aperte
if not df_positions.empty:
    df_positions['data_apertura'] = pd.to_datetime(df_positions['timestamp'], unit='s')
    cols_pos = ['id', 'data_apertura', 'market', 'side', 'entry_price', 'size', 'cost', 'leader']
    df_positions = df_positions[cols_pos]

# Creazione file Excel
with pd.ExcelWriter('trade_history_polymarket_paper.xlsx', engine='openpyxl') as writer:
    # 1. Riepilogo
    stats = data['stats']
    summary_data = {
        'Metrica': ['Saldo Iniziale', 'Saldo Attuale', 'Profitto Netto Realizzato', 'Trade Vinti', 'Trade Persi', 'Win Rate', 'ROI Totale %'],
        'Valore': [
            f"$ {data['starting_balance']}",
            f"$ {data['balance']}",
            f"$ {stats['realized_pnl']}",
            stats['wins'],
            stats['losses'],
            f"{(stats['wins'] / (stats['wins'] + stats['losses']) * 100):.2f} %",
            f"{((data['balance'] - data['starting_balance']) / data['starting_balance'] * 100):.2f} %"
        ]
    }
    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Riepilogo Generale', index=False)
    
    # 2. Storico
    df_history.to_excel(writer, sheet_name='Storico Trade (Chiusi)', index=False)
    
    # 3. Posizioni
    df_positions.to_excel(writer, sheet_name='Posizioni Aperte', index=False)

    # Formattazione estetica
    workbook = writer.book
    header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
    header_font = Font(bold=True)
    
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-fit colonne
        for column in ws.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_name].width = min(max_length + 2, 50)